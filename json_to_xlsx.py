#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
JSON to Excel Converter - Multi-Sheet Version
ORIGINAL FULL 526 LINES - FIXED: weekend duplicate + merge cells + 2 sheets
"""

import json
import re
import pandas as pd
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import copy

# 配置日志
logging.basicConfig(
    filename='schedule_fix.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)


def fix_color_code(color_code):
    if not color_code or color_code.strip() == '' or color_code == '#FFFFFF':
        return 'FFFFFF'
    color_code = color_code.replace('#', '')
    if len(color_code) == 6:
        return f'FF{color_code}'
    elif len(color_code) == 8:
        return color_code
    return 'FFFFFF'


def format_date_chinese(date_str):
    """
    将日期从 MM-DD 格式转换为 X月X日 格式
    例如: "05-07" -> "5月7日"
    """
    if '-' in str(date_str):
        parts = str(date_str).split('-')
        if len(parts) == 2:
            month = parts[0].lstrip('0')  # 去掉前导0
            day = parts[1].lstrip('0')  # 去掉前导0
            return f"{month}月{day}日"
    return date_str


def format_time_range(time_range):
    """
    将时间范围格式从 "18:00-20:00" 转换为 "18:00 至 20:00"
    """
    if '-' in str(time_range):
        parts = str(time_range).split('-')
        if len(parts) == 2:
            return f"{parts[0]} 至 {parts[1]}"
    return time_range


def fix_weekend_time_slots(time_slots, is_weekend=True):
    """
    强制修复周末时间槽格式，确保所有时间段都被拆分为独立时间点

    Args:
        time_slots: 时间槽列表，可能包含时间段格式
        is_weekend: 是否为周末（默认True）

    Returns:
        修复后的时间槽列表，确保不包含任何时间段格式
    """
    if not is_weekend:
        # 工作日不处理，允许时间段格式
        return time_slots

    fixed_slots = []
    for slot in time_slots:
        slot_str = str(slot).strip()
        if '-' in slot_str:
            # 发现时间段格式，必须拆分！
            print(f"🚨 周末发现时间段格式: {slot_str}，正在自动修复...")
            try:
                # 拆分时间段
                parts = slot_str.split('-')
                if len(parts) == 2:
                    start_time = parts[0].strip()
                    end_time = parts[1].strip()
                    # 添加开始和结束时间点
                    if start_time not in fixed_slots:
                        fixed_slots.append(start_time)
                    if end_time not in fixed_slots:
                        fixed_slots.append(end_time)
                    print(f"  ✅ 拆分结果: {start_time}, {end_time}")
                else:
                    # 无法正常拆分，保留原样
                    print(f"  ⚠️ 无法拆分，保留原样: {slot_str}")
                    if slot_str not in fixed_slots:
                        fixed_slots.append(slot_str)
            except Exception as e:
                print(f"  ❌ 拆分失败: {e}，保留原样")
                if slot_str not in fixed_slots:
                    fixed_slots.append(slot_str)
        else:
            # 已经是时间点格式，直接添加
            if slot_str not in fixed_slots:
                fixed_slots.append(slot_str)

    # 按时间排序
    def time_to_minutes(time_str):
        try:
            h, m = map(int, time_str.split(':'))
            return h * 60 + m
        except:
            return 0

    fixed_slots.sort(key=time_to_minutes)
    return fixed_slots


def format_title_date(title):
    """
    将标题中的日期格式化为MMDD格式
    例如: "应用變更人員時間安排表 2026-05-07" -> "应用變更人員時間安排表0507"
    """
    import re
    # 尝试匹配 YYYY-MM-DD 格式
    match = re.search(r'(\d{4})-(\d{2})-(\d{2})', title)
    if match:
        year = match.group(1)
        month = match.group(2)
        day = match.group(3)
        month_day = month + day  # MMDD格式，如 0507
        # 替换完整的日期字符串（包括可能的年份前缀）
        title = re.sub(r'\d{4}-\d{2}-\d{2}', month_day, title)
    else:
        # 尝试匹配 MM-DD 格式
        match = re.search(r'\d{2}-\d{2}', title)
        if match:
            month_day = match.group().replace('-', '')
            title = title.replace(match.group(), month_day)

    # 去掉"应用變更人員時間安排表"后面的空格
    title = title.replace('应用變更人員時間安排表 ', '应用變更人員時間安排表')

    return title


def create_schedule_sheet_from_latest_format(ws, header_title, structure):
    dates = structure.get('dates', [])
    participants = structure.get('staff_list', [])
    schedule_data = structure.get('schedule_data', [])
    time_slots = structure.get('time_slots', {})

    font_family = '新宋体'
    current_col = 2
    date_col_mapping = {}

    for date in dates:
        date_col_mapping[date] = []
        slots = time_slots.get(date, [])
        if any('-' in slot for slot in slots):
            date_col_mapping[date].append(current_col)
            ws.column_dimensions[get_column_letter(current_col)].width = 28
            current_col += 1
        else:
            for _ in slots:
                date_col_mapping[date].append(current_col)
                ws.column_dimensions[get_column_letter(current_col)].width = 25
                current_col += 1

    ws.column_dimensions['A'].width = 33
    total_cols = current_col - 1

    ws.row_dimensions[1].height = 45
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = header_title
    title_cell.font = Font(name=font_family, size=30, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_fill_color = fix_color_code('#E6F3FF')
    title_cell.fill = PatternFill(start_color=title_fill_color, end_color=title_fill_color, fill_type='solid')
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    ws.row_dimensions[2].height = 35
    date_fill_color = fix_color_code('#E6F3FF')
    for date in dates:
        cols = date_col_mapping[date]
        if not cols:
            continue
        start_c = cols[0]
        end_c = cols[-1]
        cell = ws.cell(row=2, column=start_c)
        cell.value = date
        cell.font = Font(name=font_family, size=26, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=date_fill_color, end_color=date_fill_color, fill_type='solid')
        if start_c != end_c:
            ws.merge_cells(start_row=2, start_column=start_c, end_row=2, end_column=end_c)

    ws.row_dimensions[3].height = 30
    for date in dates:
        cols = date_col_mapping[date]
        slots = time_slots.get(date, [])
        for i, slot in enumerate(slots):
            if i >= len(cols):
                break
            cell = ws.cell(row=3, column=cols[i])
            cell.value = slot
            cell.font = Font(name=font_family, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # ====================== 修复：去重 + 合并单元格 ======================
    for row_idx, staff_data in enumerate(schedule_data, start=4):
        person_name = staff_data.get('staff', '')
        color = staff_data.get('color', '#FFFFFF')
        tasks = staff_data.get('tasks', {})
        ws.row_dimensions[row_idx].height = 90
        bg_color = fix_color_code(color)

        person_cell = ws.cell(row=row_idx, column=1)
        person_cell.value = person_name
        person_cell.font = Font(name=font_family, size=14, bold=True)
        person_cell.alignment = Alignment(horizontal='center', vertical='center')
        person_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        for date in dates:
            date_task_list = tasks.get(date, [])
            if not date_task_list:
                continue

            slot_list = time_slots.get(date, [])
            col_list = date_col_mapping.get(date, [])
            if not slot_list or not col_list:
                continue

            unique_tasks = []
            seen = set()
            for task_items in date_task_list:
                task_str = "\n".join(task_items)
                if task_str not in seen:
                    seen.add(task_str)
                    unique_tasks.append(task_items)

            for task_items in unique_tasks:
                if not task_items:
                    continue

                min_col = col_list[0]
                max_col = col_list[-1]
                task_text = "\n".join(task_items)

                if min_col != max_col:
                    ws.merge_cells(start_row=row_idx, start_column=min_col, end_row=row_idx, end_column=max_col)

                task_cell = ws.cell(row=row_idx, column=min_col)
                task_cell.value = task_text
                task_cell.font = Font(name=font_family, size=12)
                task_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                task_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=len(schedule_data)+3, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = thin_border


def create_schedule_sheet_from_headers_format(ws, structure):
    """
    处理最新的headers格式
    structure: {
        'title': '应用變更人員時間安排表05-07',
        'headers': {
            'row_2_dates': ['05-04', '05-05', ...],
            'row_3_times': {'05-04': ['18:00-19:00'], '05-09': ['12:00', '12:30'], ...}
        },
        'personnel_rows': [
            {
                'name': '李俊毅',
                'color': '#E8F5E8',
                'assignments': [
                    {'date': '05-07', 'time_range': '18:00-19:00', 'work_order': '...'}
                ]
            },
            ...
        ]
    }
    """
    font_family = '新宋体'
    title = structure.get('title', '应用變更人員時間安排表')
    headers = structure.get('headers', {})
    personnel_rows = structure.get('personnel_rows', [])

    dates = headers.get('row_2_dates', [])
    time_slots = headers.get('row_3_times', {})

    # Issue 6: 标题格式修正
    title = format_title_date(title)

    current_col = 2
    date_col_mapping = {}

    # 为每个日期分配列
    for date in dates:
        date_col_mapping[date] = []
        slots = time_slots.get(date, [])

        # 🚨 检查是否为周末（通过时间槽格式判断）
        has_time_range = any('-' in slot for slot in slots)

        # 如果有时间段格式且是周末，强制修复
        if has_time_range:
            # 检查是否为周末（通过日期字符串判断或其他方式）
            # 这里我们假设如果有时间段且看起来像周末，先尝试修复
            print(f"检查日期 {date} 的时间槽: {slots}")
            fixed_slots = fix_weekend_time_slots(slots, is_weekend=True)

            # 如果修复后时间槽数量不同，说明确实是周末需要修复
            if len(fixed_slots) != len(slots):
                print(f"✅ 修复周末时间槽: {date}")
                print(f"  原始: {slots}")
                print(f"  修复: {fixed_slots}")
                time_slots[date] = fixed_slots
                slots = fixed_slots
                # 重新判断
                has_time_range = any('-' in slot for slot in slots)

        # 🔧 补全周末时间槽：确保包含所有工单中提到的时间点
        if not has_time_range:  # 周末（时间点格式）
            required_times = set(slots)
            # 从personnel_rows中收集该日期的所有工单时间
            for person_data in personnel_rows:
                for assignment in person_data.get('assignments', []):
                    if assignment.get('date') == date:
                        start_time = assignment.get('start_time', '')
                        end_time = assignment.get('end_time', '')
                        time_range = assignment.get('time_range', '')

                        # 提取时间点
                        if start_time:
                            required_times.add(start_time.split('-')[0] if '-' in start_time else start_time)
                        if end_time:
                            required_times.add(end_time.split('-')[0] if '-' in end_time else end_time)
                        if time_range and '-' in time_range:
                            start, end = time_range.split('-')
                            required_times.add(start)
                            required_times.add(end)

            # 按时间排序
            def time_to_minutes(time_str):
                try:
                    h, m = map(int, time_str.split(':'))
                    return h * 60 + m
                except:
                    return 0

            required_slots = sorted(list(required_times), key=time_to_minutes)
            if len(required_slots) != len(slots):
                print(f"🔧 补全周末时间槽: {date}")
                print(f"  原始: {slots}")
                print(f"  补全: {required_slots}")
                time_slots[date] = required_slots
                slots = required_slots

        if has_time_range:
            # 时间段：占一列（工作日）
            date_col_mapping[date].append(current_col)
            # Issue 4: 工作日列宽32
            ws.column_dimensions[get_column_letter(current_col)].width = 32
            current_col += 1
        else:
            # 时间点：每个时间点占一列（周末）
            for _ in slots:
                date_col_mapping[date].append(current_col)
                # Issue 4: 周末列宽16
                ws.column_dimensions[get_column_letter(current_col)].width = 16
                current_col += 1

    ws.column_dimensions['A'].width = 33
    total_cols = current_col - 1

    # 标题行
    ws.row_dimensions[1].height = 45
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(name=font_family, size=30, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_fill_color = fix_color_code('#E6F3FF')
    title_cell.fill = PatternFill(start_color=title_fill_color, end_color=title_fill_color, fill_type='solid')
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Issue 2: A2和A3合并，内容寫執行人員
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    person_label_cell = ws.cell(row=2, column=1)
    person_label_cell.value = '執行人員'
    person_label_cell.font = Font(name=font_family, size=14, bold=True)
    person_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    person_label_cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')

    # 日期行
    ws.row_dimensions[2].height = 35
    date_fill_color = fix_color_code('#E6F3FF')
    for date in dates:
        cols = date_col_mapping.get(date, [])
        if not cols:
            continue
        start_c = cols[0]
        end_c = cols[-1]
        cell = ws.cell(row=2, column=start_c)
        # Issue 3: 日期格式从MM-DD改为X月X日
        cell.value = format_date_chinese(date)
        cell.font = Font(name=font_family, size=26, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=date_fill_color, end_color=date_fill_color, fill_type='solid')
        if start_c != end_c:
            ws.merge_cells(start_row=2, start_column=start_c, end_row=2, end_column=end_c)

    # 时间行
    ws.row_dimensions[3].height = 30
    for date in dates:
        cols = date_col_mapping.get(date, [])
        slots = time_slots.get(date, [])
        for i, slot in enumerate(slots):
            if i >= len(cols):
                break
            cell = ws.cell(row=3, column=cols[i])
            cell.value = slot
            cell.font = Font(name=font_family, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 人员行
    for row_idx, person_data in enumerate(personnel_rows, start=4):
        person_name = person_data.get('name', '')
        color = person_data.get('color', '#FFFFFF')
        assignments = person_data.get('assignments', [])
        # Issue 5: 第4行以后的行行高80
        ws.row_dimensions[row_idx].height = 80
        bg_color = fix_color_code(color)

        # 姓名单元格
        person_cell = ws.cell(row_idx, column=1)
        person_cell.value = person_name
        person_cell.font = Font(name=font_family, size=14, bold=True)
        person_cell.alignment = Alignment(horizontal='center', vertical='center')
        person_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        # 工单单元格（支持跨时间点合并）
        for assignment in assignments:
            date = assignment.get('date', '')
            start_time = assignment.get('start_time', '')
            end_time = assignment.get('end_time', '')
            time_range = assignment.get('time_range', '')
            work_order = assignment.get('work_order', '')

            if date in date_col_mapping:
                cols = date_col_mapping[date]
                slots = time_slots.get(date, [])

                # 检查是否跨时间点
                is_cross_slot = False
                target_cols = []

                # 判断slots是否为时间范围（工作日）还是时间点（周末）
                has_time_range_slots = any('-' in str(slot) for slot in slots)

                if start_time and end_time and start_time != end_time:
                    # 将时间字符串转换为分钟数进行比较
                    def time_to_minutes(time_str):
                        try:
                            # Handle dashed time ranges like '18:00-19:00' by extracting start time
                            if '-' in str(time_str):
                                time_str = time_str.split('-')[0]
                            h, m = map(int, time_str.split(':'))
                            return h * 60 + m
                        except:
                            return -1

                    start_minutes = time_to_minutes(start_time)
                    end_minutes = time_to_minutes(end_time)

                    if has_time_range_slots:
                        # 工作日：slots是时间范围，匹配包含该时间范围的列
                        for i, slot in enumerate(slots):
                            if '-' in str(slot):
                                # 提取slot的时间范围
                                try:
                                    slot_start, slot_end = str(slot).split('-')
                                    slot_start_min = time_to_minutes(slot_start.strip())
                                    slot_end_min = time_to_minutes(slot_end.strip())
                                    # 检查是否有重叠
                                    if not (end_minutes < slot_start_min or start_minutes > slot_end_min):
                                        if i < len(cols):
                                            target_cols.append(cols[i])
                                            is_cross_slot = True
                                except:
                                    # 如果解析失败，检查字符串匹配
                                    if str(slot) in time_range or time_range in str(slot):
                                        if i < len(cols):
                                            target_cols.append(cols[i])
                                            is_cross_slot = True
                        target_cols = list(sorted(set(target_cols)))  # 去重
                    else:
                        # 周末：slots是时间点，检查跨越哪些时间点
                        for i, slot in enumerate(slots):
                            slot_minutes = time_to_minutes(slot)
                            # 检查时间点是否在范围内（包含边界）
                            if start_minutes <= slot_minutes <= end_minutes:
                                if i < len(cols):
                                    target_cols.append(cols[i])
                                    is_cross_slot = True
                        target_cols = list(sorted(set(target_cols)))  # 去重
                elif time_range and len(cols) > 0:
                    # 没有明确的start/end时间，使用time_range匹配
                    if has_time_range_slots:
                        # 工作日：匹配时间范围
                        for i, slot in enumerate(slots):
                            if str(slot) == time_range or (time_range in str(slot) and '-' in str(slot)):
                                if i < len(cols):
                                    target_cols.append(cols[i])
                                    break  # 工作日只匹配一列
                    else:
                        # 周末：尝试匹配多个时间点
                        for i, slot in enumerate(slots):
                            if ('-' in time_range and slot in time_range) or slot == time_range:
                                if i < len(cols):
                                    target_cols.append(cols[i])
                        target_cols = list(sorted(set(target_cols)))
                        if len(target_cols) > 1:
                            is_cross_slot = True

                # 如果还是没有匹配到，默认使用第一列
                if not target_cols and cols:
                    target_cols = [cols[0]]

                # 合并单元格（如果跨时间点）
                if is_cross_slot and len(target_cols) > 1:
                    min_col = min(target_cols)
                    max_col = max(target_cols)

                    # 合并单元格
                    ws.merge_cells(start_row=row_idx, start_column=min_col, end_row=row_idx, end_column=max_col)

                    # 放置工单内容（在第一个单元格）
                    cell = ws.cell(row_idx, column=min_col)
                    if work_order:
                        cell.value = work_order
                        cell.font = Font(name=font_family, size=12)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                elif target_cols and work_order:
                    # 单个时间点，不合并
                    target_col = target_cols[0]
                    cell = ws.cell(row_idx, column=target_col)
                    cell.value = work_order
                    cell.font = Font(name=font_family, size=12)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=len(personnel_rows)+3, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = thin_border


def create_schedule_sheet_from_new_format(ws, header_title, participants, schedule_data):
    columns = schedule_data['columns']
    rows = schedule_data['rows']
    dates = []
    for col in columns[1:]:
        if col['type'] == 'date':
            dates.append(col['content'])

    font_family = '新宋体'
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = header_title
    title_cell.font = Font(name=font_family, size=30, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_fill_color = fix_color_code('#E6F3FF')
    title_cell.fill = PatternFill(start_color=title_fill_color, end_color=title_fill_color, fill_type='solid')
    total_cols = len(dates) + 1
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    ws.row_dimensions[2].height = 30
    date_fill_color = fix_color_code('#E6F3FF')
    for col_idx, date in enumerate(dates, start=2):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = date
        cell.font = Font(name=font_family, size=26, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=date_fill_color, end_color=date_fill_color, fill_type='solid')

    ws.row_dimensions[3].height = 25
    for col_idx, date in enumerate(dates, start=2):
        cell = ws.cell(row=3, column=col_idx)
        time_points = []
        for row in rows:
            tasks = row.get('tasks', {}).get(date, [])
            for task in tasks:
                if 'time_range' in task:
                    time_points.append(task['time_range'])
                elif 'time_point' in task:
                    time_points.append(task['time_point'])
        cell.value = '/'.join(list(set(time_points))) if time_points else date
        cell.font = Font(name=font_family, size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row_data in enumerate(rows, start=4):
        person_name = row_data.get('person', '')
        color = row_data.get('color', '#FFFFFF')
        tasks = row_data.get('tasks', {})
        ws.row_dimensions[row_idx].height = 90

        person_cell = ws.cell(row=row_idx, column=1)
        person_cell.value = person_name
        person_cell.font = Font(name=font_family, size=14, bold=True)
        person_cell.alignment = Alignment(horizontal='center', vertical='center')
        person_cell.fill = PatternFill(start_color=fix_color_code(color), end_color=fix_color_code(color), fill_type='solid')

        for col_idx, date in enumerate(dates, start=2):
            date_tasks = tasks.get(date, [])
            display_texts = [t['display_text'] for t in date_tasks if 'display_text' in t]
            task_cell = ws.cell(row=row_idx, column=col_idx)
            task_cell.value = '\n'.join(display_texts)
            task_cell.font = Font(name=font_family, size=14)
            task_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if display_texts:
                task_cell.fill = PatternFill(start_color=fix_color_code(color), end_color=fix_color_code(color), fill_type='solid')

    ws.column_dimensions['A'].width = 33
    for col_idx in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    for row in ws.iter_rows(min_row=1, max_row=len(rows)+3, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = thin_border


def create_schedule_sheet_from_assignments_format(ws, data):
    """
    处理assignments格式 - 完全泛化的实现
    严格按照columns定义创建列结构，智能判断工作日/周末
    """
    font_family = '新宋体'
    sheet_structure = data.get('sheet_structure', {})
    assignments = data.get('assignments', [])

    title = sheet_structure.get('title_row', '应用變更人員時間安排表')
    columns = sheet_structure.get('columns', [])
    time_row_data = sheet_structure.get('time_row_data', {})

    # Issue 6: 标题格式修正
    title = format_title_date(title)

    # 严格按照columns定义创建列结构
    dates = []
    date_col_mapping = {}
    current_col = 2

    # 按日期分组columns（保持原始顺序）
    date_columns = {}  # {date: [col_names]}
    for col in columns[1:]:  # 跳过第一列"执行人"
        if '_' in col:
            date_part = col.split('_')[0]
        else:
            date_part = col

        if date_part not in date_columns:
            date_columns[date_part] = []
        date_columns[date_part].append(col)

    # 按columns中第一次出现的顺序创建列
    seen_dates = []
    for col in columns[1:]:
        if '_' in col:
            date_part = col.split('_')[0]
        else:
            date_part = col
        if date_part not in seen_dates:
            seen_dates.append(date_part)

    # 分析每个日期是工作日还是周末
    date_is_weekend = {}

    # 🚨 首先检查并修复所有周末时间槽格式
    print("=== 开始检查周末时间槽格式 (assignments格式) ===")
    for assignment in assignments:
        for task in assignment.get('tasks', []):
            date = task.get('date', '')
            if date:
                time_slots = time_row_data.get(date, [])
                # 检查是否包含时间段格式
                has_time_range = any('-' in str(slot) for slot in time_slots)

                if has_time_range:
                    # 可能是周末的时间段格式，需要修复
                    print(f"检查日期 {date} 的时间槽: {time_slots}")
                    fixed_slots = fix_weekend_time_slots(time_slots, is_weekend=True)

                    if len(fixed_slots) != len(time_slots) or fixed_slots != time_slots:
                        print(f"✅ 修复周末时间槽: {date}")
                        print(f"  原始: {time_slots}")
                        print(f"  修复: {fixed_slots}")
                        time_row_data[date] = fixed_slots
                        time_slots = fixed_slots

                # 判断是否为工作日
                is_workday = (len(time_slots) > 0 and
                             all(slot == time_slots[0] for slot in time_slots) and
                             '-' not in str(time_slots[0]))
                if date not in date_is_weekend:
                    date_is_weekend[date] = not is_workday
    print("=== 周末时间槽格式检查完成 ===")

    for date in seen_dates:
        date_col_mapping[date] = []
        for col in date_columns[date]:
            date_col_mapping[date].append(current_col)
            # Issue 4: 根据工作日/周末设置不同列宽
            if date_is_weekend.get(date, False):
                ws.column_dimensions[get_column_letter(current_col)].width = 16  # 周末列宽16
            else:
                ws.column_dimensions[get_column_letter(current_col)].width = 32  # 工作日列宽32
            current_col += 1
        if date not in dates:
            dates.append(date)

    ws.column_dimensions['A'].width = 33
    total_cols = current_col - 1

    # 标题行
    ws.row_dimensions[1].height = 45
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(name=font_family, size=30, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_fill_color = fix_color_code('#E6F3FF')
    title_cell.fill = PatternFill(start_color=title_fill_color, end_color=title_fill_color, fill_type='solid')
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Issue 2: A2和A3合并，内容寫執行人員
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    person_label_cell = ws.cell(row=2, column=1)
    person_label_cell.value = '執行人員'
    person_label_cell.font = Font(name=font_family, size=14, bold=True)
    person_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    person_label_cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')

    # 日期行
    ws.row_dimensions[2].height = 35
    date_fill_color = fix_color_code('#E6F3FF')
    for date in dates:
        cols = date_col_mapping.get(date, [])
        if not cols:
            continue
        start_c = cols[0]
        end_c = cols[-1]
        cell = ws.cell(row=2, column=start_c)
        # Issue 3: 日期格式从MM-DD改为X月X日
        cell.value = format_date_chinese(date)
        cell.font = Font(name=font_family, size=26, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=date_fill_color, end_color=date_fill_color, fill_type='solid')
        if start_c != end_c:
            ws.merge_cells(start_row=2, start_column=start_c, end_row=2, end_column=end_c)

    # 时间行（智能判断工作日/周末）
    ws.row_dimensions[3].height = 30
    for date in dates:
        cols = date_col_mapping.get(date, [])
        time_slots = time_row_data.get(date, [])

        # 判断是否为工作日：所有时间槽相同且不包含'-'
        is_workday = (len(time_slots) > 0 and
                     all(slot == time_slots[0] for slot in time_slots) and
                     '-' not in str(time_slots[0]))

        if is_workday:
            # 工作日：查找最长的时间范围
            best_start = None
            best_end = None
            best_duration = 0

            for assignment in assignments:
                for task in assignment.get('tasks', []):
                    if task.get('date') == date:
                        start = task.get('original_start', '')
                        end = task.get('original_end', '')
                        if start and end:
                            try:
                                start_min = int(start.split(':')[0]) * 60 + int(start.split(':')[1])
                                end_min = int(end.split(':')[0]) * 60 + int(end.split(':')[1])
                                duration = end_min - start_min
                                if duration > best_duration:
                                    best_duration = duration
                                    best_start = start
                                    best_end = end
                            except:
                                pass

            time_range = f"{best_start}-{best_end}" if best_start and best_end else time_slots[0]

            # 只在第一列显示时间范围
            for i, col in enumerate(cols):
                cell = ws.cell(row=3, column=col)
                if i == 0:
                    cell.value = time_range
                else:
                    cell.value = ''  # 其他列为空
                cell.font = Font(name=font_family, size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            # 周末：按列显示时间点
            # 🚨 再次检查并确保时间点格式正确
            print(f"显示周末时间行: {date}, 时间槽: {time_slots}")
            for i, time_slot in enumerate(time_slots):
                if i >= len(cols):
                    break
                cell = ws.cell(row=3, column=cols[i])
                # 确保不显示时间段格式
                slot_display = str(time_slot).strip()
                if '-' in slot_display and len(slot_display) > 5:
                    # 如果仍然是时间段格式，尝试只显示开始时间
                    print(f"⚠️ 警告：周末时间槽仍然包含时间段格式: {slot_display}")
                    slot_display = slot_display.split('-')[0].strip()
                cell.value = slot_display
                cell.font = Font(name=font_family, size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 人员行（正确分配工单）
    for row_idx, assignment in enumerate(assignments, start=4):
        person_name = assignment.get('personnel', '')
        color = assignment.get('color', '#FFFFFF')
        tasks = assignment.get('tasks', [])
        # Issue 5: 第4行以后的行行高80
        ws.row_dimensions[row_idx].height = 80
        bg_color = fix_color_code(color)

        # 姓名单元格
        person_cell = ws.cell(row_idx, column=1)
        person_cell.value = person_name
        person_cell.font = Font(name=font_family, size=14, bold=True)
        person_cell.alignment = Alignment(horizontal='center', vertical='center')
        person_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        # 任务单元格
        for task in tasks:
            date = task.get('date', '')
            work_order = task.get('work_order', '')
            original_start = task.get('original_start', '')
            original_end = task.get('original_end', '')
            merge_columns = task.get('merge_columns', [])

            if date not in date_col_mapping:
                continue

            cols = date_col_mapping[date]
            time_slots = time_row_data.get(date, [])

            # 判断是否为工作日
            is_workday = (len(time_slots) > 0 and
                         all(slot == time_slots[0] for slot in time_slots) and
                         '-' not in str(time_slots[0]))

            if is_workday:
                # 工作日：使用第一列
                target_col = cols[0]
                cell = ws.cell(row_idx, column=target_col)
                try:
                    current_value = cell.value if cell.value else ""
                    if current_value and str(current_value).strip():
                        # 如果已经有内容，添加换行
                        cell.value = str(current_value) + '\n' + work_order
                    else:
                        cell.value = work_order
                    cell.font = Font(name=font_family, size=12)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                except Exception as e:
                    print(f"⚠️ 警告：工作日写入单元格时出错: {e}")
                    # 如果写入失败，直接覆盖
                    try:
                        cell.value = work_order
                        cell.font = Font(name=font_family, size=12)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                    except:
                        print(f"❌ 错误：无法写入工单到工作日单元格")
                        raise
            else:
                # 周末：使用merge_columns或匹配时间点
                target_cols = []

                if merge_columns:
                    # 使用merge_columns确定目标列
                    for merge_col in merge_columns:
                        if '_' in merge_col:
                            date_part = merge_col.split('_')[0]
                            slot_num = int(merge_col.split('_')[1]) - 1
                            if date_part in date_col_mapping:
                                cols_for_date = date_col_mapping[date_part]
                                if slot_num < len(cols_for_date):
                                    target_cols.append(cols_for_date[slot_num])
                else:
                    # 通过时间匹配找到目标列
                    start_time = task.get('original_start', '')
                    end_time = task.get('original_end', '')

                    if start_time and end_time:
                        # 转换时间为分钟
                        def time_to_minutes(time_str):
                            try:
                                # Handle dashed time ranges like '18:00-19:00' by extracting start time
                                if '-' in str(time_str):
                                    time_str = time_str.split('-')[0]
                                h, m = map(int, time_str.split(':'))
                                return h * 60 + m
                            except:
                                return -1

                        start_minutes = time_to_minutes(start_time)
                        end_minutes = time_to_minutes(end_time)

                        # 检查跨越哪些时间点
                        for i, slot in enumerate(time_slots):
                            slot_minutes = time_to_minutes(slot)
                            if start_minutes <= slot_minutes <= end_minutes:
                                if i < len(cols):
                                    target_cols.append(cols[i])

                # 合并单元格并放置工单
                if target_cols:
                    target_cols = sorted(list(set(target_cols)))
                    if len(target_cols) > 1:
                        min_col = min(target_cols)
                        max_col = max(target_cols)
                        ws.merge_cells(start_row=row_idx, start_column=min_col, end_row=row_idx, end_column=max_col)
                        cell = ws.cell(row_idx, column=min_col)
                    else:
                        cell = ws.cell(row_idx, column=target_cols[0])

                    if work_order:
                        cell.value = work_order
                        cell.font = Font(name=font_family, size=12)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=len(assignments)+3, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = thin_border



def load_filtered_excel_data(excel_file='data/sheet_input.xlsx'):
    if not os.path.exists(excel_file):
        return None
    df = pd.read_excel(excel_file)
    if '工单状态' in df.columns:
        df = df[df['工单状态'] != '已撤銷']
    filter_teams = ['安全內控團隊', '前端網絡團隊', '系統平臺團隊']
    if '提單人所屬團隊' in df.columns:
        df = df[~df['提單人所屬團隊'].isin(filter_teams)]
    if '序号' in df.columns:
        df = df.drop(columns=['序号'])
    return df


def create_data_sheet_from_dataframe(ws, df):
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')

    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row_data[col_name]
            cell.font = Font(name='Arial', size=10)
            # Issue 1: 變更明細數據報表内容不要換行或者固定行高
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    thin_border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    for row in ws.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.border = thin_border


def ensure_data_completeness(json_data, original_excel_path='data/sheet_input.xlsx'):
    logging.info("数据完整性检查...")
    if not os.path.exists(original_excel_path):
        return json_data
    original_df = pd.read_excel(original_excel_path)
    original_columns = set(original_df.columns) - {'序号'}
    return json_data


def split_weekend_time_slots(schedule_data):
    logging.info("周末时间不拆分，使用原始格式")
    return schedule_data


def load_json_result(json_file='ai_responses/schedule_result.json'):
    with open(json_file, 'r', encoding='utf-8') as f:
        content = f.read()
    match = re.search(r'```json\s*(.*?)\s*```', content, re.DOTALL)
    if match:
        content = match.group(1).strip()
    return json.loads(content.strip())


def create_schedule_sheet_from_new_structure_format(ws, data):
    """
    处理新的JSON格式 - schedule_structure + personnel_assignments
    data: {
        'schedule_structure': {
            'header_title': '应用變更人員時間安排表 05-07',
            'dates': ['05-04', '05-05', '05-07', '05-09', '05-10'],
            'time_slots_by_date': {
                '05-04': ['18:00-19:00'],
                '05-05': ['09:00-10:00'],
                '05-07': ['18:00-20:00'],
                '05-09': ['12:00', '12:30'],
                '05-10': ['01:00', '02:00', '03:00', '06:00', '10:00', '11:00']
            }
        },
        'personnel_assignments': [
            {
                'name': '李俊毅',
                'color_code': '#E8F5E8',
                'assignments': [
                    {
                        'date': '05-07',
                        'time_slot_index': 0,
                        'work_order': '(中台) 02237 TBU 王文斌',
                        'original_start': '18:00',
                        'original_end': '19:00',
                        'is_weekend': false
                    },
                    ...
                ]
            },
            ...
        ]
    }
    """
    font_family = '新宋体'
    schedule_structure = data.get('schedule_structure', {})
    personnel_assignments = data.get('personnel_assignments', [])

    title = schedule_structure.get('header_title', '应用變更人員時間安排表')
    dates = schedule_structure.get('dates', [])
    time_slots_by_date = schedule_structure.get('time_slots_by_date', {})

    # Issue 6: 标题格式修正 - 去掉多余的"-"，改为"应用變更人員時間安排表 MMDD"
    title = format_title_date(title)

    # 创建列映射
    date_col_mapping = {}
    current_col = 2

    # 首先分析每个日期是工作日还是周末
    date_is_weekend = {}  # {date: is_weekend}

    for person_data in personnel_assignments:
        for assignment in person_data.get('assignments', []):
            date = assignment.get('date', '')
            is_weekend = assignment.get('is_weekend', False)
            if date and date not in date_is_weekend:
                date_is_weekend[date] = is_weekend

    # 🚨 强制检查并修复所有周末时间槽格式
    print("=== 开始检查周末时间槽格式 ===")
    for date in dates:
        if date_is_weekend.get(date, False):
            time_slots = time_slots_by_date.get(date, [])
            print(f"检查周末日期: {date}, 时间槽: {time_slots}")
            fixed_slots = fix_weekend_time_slots(time_slots, is_weekend=True)
            if len(fixed_slots) != len(time_slots) or fixed_slots != time_slots:
                print(f"✅ 修复周末时间槽: {date}")
                print(f"  原始: {time_slots}")
                print(f"  修复: {fixed_slots}")
                time_slots_by_date[date] = fixed_slots
    print("=== 周末时间槽格式检查完成 ===")

    # 🔧 新增：确保周末时间槽包含所有工单中提到的时间点
    print("=== 开始补全周末时间槽 ===")
    weekend_slot_mapping = {}  # 记录时间槽变化，用于后续索引调整

    for date in dates:
        if date_is_weekend.get(date, False):
            time_slots = time_slots_by_date.get(date, [])
            original_slots = list(time_slots)  # 保存原始时间槽
            required_times = set(time_slots)  # 从现有时间槽开始

            # 遍历所有人员在该日期的工单，收集所有需要的时间点
            for person_data in personnel_assignments:
                for assignment in person_data.get('assignments', []):
                    if assignment.get('date') == date:
                        original_start = assignment.get('original_start', '')
                        original_end = assignment.get('original_end', '')

                        if original_start:
                            required_times.add(original_start)
                        if original_end:
                            required_times.add(original_end)

            # 转换为分钟数进行排序
            def time_to_minutes(time_str):
                try:
                    h, m = map(int, time_str.split(':'))
                    return h * 60 + m
                except:
                    return 0

            required_times_list = sorted(list(required_times), key=time_to_minutes)

            # 如果发现了新的时间点，更新time_slots
            if len(required_times_list) != len(time_slots) or required_times_list != time_slots:
                print(f"🔧 补全周末时间槽: {date}")
                print(f"  原始: {time_slots}")
                print(f"  补全: {required_times_list}")
                time_slots_by_date[date] = required_times_list
                # 记录映射关系
                weekend_slot_mapping[date] = {
                    'original': original_slots,
                    'updated': required_times_list
                }

    print("=== 周末时间槽补全完成 ===")
    if weekend_slot_mapping:
        print("⚠️ 警告：周末时间槽已补全，合并索引将重新计算")
        print(f"影响的日期: {list(weekend_slot_mapping.keys())}")

    for date in dates:
        time_slots = time_slots_by_date.get(date, [])
        num_slots = len(time_slots)

        if date not in date_col_mapping:
            date_col_mapping[date] = []

        # 判断是否为工作日
        is_workday = not date_is_weekend.get(date, False)

        if is_workday:
            # 工作日：只创建一列
            date_col_mapping[date].append(current_col)
            # Issue 4: 工作日列宽32
            ws.column_dimensions[get_column_letter(current_col)].width = 32
            current_col += 1
        else:
            # 周末：收集所有工单的开始时间和结束时间，确保所有时间点都存在
            def time_to_minutes(time_str):
                try:
                    # Handle dashed time ranges like '18:00-19:00' by extracting start time
                    if '-' in str(time_str):
                        time_str = time_str.split('-')[0]
                    h, m = map(int, time_str.split(':'))
                    return h * 60 + m
                except:
                    return -1

            # 收集所有需要的时间点
            required_times = set(time_slots)  # 从原始时间槽开始

            # 添加所有工单的开始时间和结束时间
            for person_data in personnel_assignments:
                for assignment in person_data.get('assignments', []):
                    if assignment.get('date') == date:
                        original_start = assignment.get('original_start', '')
                        original_end = assignment.get('original_end', '')
                        if original_start:
                            required_times.add(original_start)
                        if original_end:
                            required_times.add(original_end)

            # 转换为列表并按时间排序
            required_slots = sorted(list(required_times), key=time_to_minutes)

            # 为每个需要的时间槽创建列
            for _ in required_slots:
                date_col_mapping[date].append(current_col)
                # Issue 4: 周末列宽16
                ws.column_dimensions[get_column_letter(current_col)].width = 16
                current_col += 1

            # 更新time_slots_by_date以包含所有需要的时间点
            time_slots_by_date[date] = required_slots

    ws.column_dimensions['A'].width = 33
    total_cols = current_col - 1

    # 标题行
    ws.row_dimensions[1].height = 45
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(name=font_family, size=30, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_fill_color = fix_color_code('#E6F3FF')
    title_cell.fill = PatternFill(start_color=title_fill_color, end_color=title_fill_color, fill_type='solid')
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Issue 2: A2和A3合并，内容寫執行人員
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    person_label_cell = ws.cell(row=2, column=1)
    person_label_cell.value = '執行人員'
    person_label_cell.font = Font(name=font_family, size=14, bold=True)
    person_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    person_label_cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')

    # 日期行
    ws.row_dimensions[2].height = 35
    date_fill_color = fix_color_code('#E6F3FF')
    for date in dates:
        cols = date_col_mapping.get(date, [])
        if not cols:
            continue
        start_c = cols[0]
        end_c = cols[-1]
        cell = ws.cell(row=2, column=start_c)
        # Issue 3: 日期格式从MM-DD改为X月X日
        cell.value = format_date_chinese(date)
        cell.font = Font(name=font_family, size=26, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=date_fill_color, end_color=date_fill_color, fill_type='solid')
        if start_c != end_c:
            ws.merge_cells(start_row=2, start_column=start_c, end_row=2, end_column=end_c)

    # 时间行
    ws.row_dimensions[3].height = 30
    for date in dates:
        cols = date_col_mapping.get(date, [])
        time_slots = time_slots_by_date.get(date, [])

        # 判断是否为工作日
        is_workday = not date_is_weekend.get(date, False)

        if is_workday:
            # 工作日：需要从工单中提取最长的时间范围来显示
            # 查找该日期所有工单的时间范围
            best_start = None
            best_end = None
            best_duration = 0

            for person_data in personnel_assignments:
                for task in person_data.get('assignments', []):
                    if task.get('date') == date:
                        start = task.get('original_start', '')
                        end = task.get('original_end', '')
                        if start and end:
                            try:
                                start_min = int(start.split(':')[0]) * 60 + int(start.split(':')[1])
                                end_min = int(end.split(':')[0]) * 60 + int(end.split(':')[1])
                                duration = end_min - start_min
                                if duration > best_duration:
                                    best_duration = duration
                                    best_start = start
                                    best_end = end
                            except:
                                pass

            # 如果找到了时间范围，显示它；否则显示原始时间槽
            if best_start and best_end:
                formatted_time = f"{best_start} 至 {best_end}"
            elif time_slots:
                # 使用原始时间槽，尝试格式化
                if '-' in str(time_slots[0]):
                    formatted_time = format_time_range(time_slots[0])
                else:
                    formatted_time = time_slots[0]
            else:
                formatted_time = "时间待定"

            for col in cols:
                cell = ws.cell(row=3, column=col)
                cell.value = formatted_time
                cell.font = Font(name=font_family, size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            # 周末：显示时间点
            # 🚨 强制检查并修复周末时间槽格式
            print(f"检查周末时间槽: {date} - {time_slots}")
            fixed_time_slots = fix_weekend_time_slots(time_slots, is_weekend=True)

            # 如果修复后时间槽数量不同，更新time_slots_by_date
            if len(fixed_time_slots) != len(time_slots):
                print(f"✅ 周末时间槽已修复: {date}")
                print(f"  原始: {time_slots}")
                print(f"  修复: {fixed_time_slots}")
                time_slots_by_date[date] = fixed_time_slots
                time_slots = fixed_time_slots

            # 显示修复后的时间点
            for i, time_slot in enumerate(time_slots):
                if i >= len(cols):
                    break
                cell = ws.cell(row=3, column=cols[i])
                cell.value = time_slot
                cell.font = Font(name=font_family, size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 人员行
    for row_idx, person_data in enumerate(personnel_assignments, start=4):
        person_name = person_data.get('name', '')
        color_code = person_data.get('color_code', '#FFFFFF')
        assignments = person_data.get('assignments', [])
        # Issue 5: 第4行以后的行行高80
        ws.row_dimensions[row_idx].height = 80
        bg_color = fix_color_code(color_code)

        # 姓名单元格
        person_cell = ws.cell(row_idx, column=1)
        person_cell.value = person_name
        person_cell.font = Font(name=font_family, size=14, bold=True)
        person_cell.alignment = Alignment(horizontal='center', vertical='center')
        person_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        # 任务单元格
        for assignment in assignments:
            date = assignment.get('date', '')
            time_slot_index = assignment.get('time_slot_index', 0)
            work_order = assignment.get('work_order', '')
            original_start = assignment.get('original_start', '')
            original_end = assignment.get('original_end', '')
            merge_info = assignment.get('merge_info', {})

            if date not in date_col_mapping:
                continue

            cols = date_col_mapping[date]
            time_slots = time_slots_by_date.get(date, [])

            # 判断是否需要合并单元格
            should_merge = False
            start_slot = time_slot_index
            end_slot = time_slot_index

            # 判断是否为周末
            is_weekend = date_is_weekend.get(date, False)

            # 🔧 周末工单合并逻辑：完全忽略AI的索引，重新计算
            if is_weekend and original_start and original_end and time_slots:
                # 周末：自动判断是否需要合并：查找original_start和original_end对应的时间点
                def time_to_minutes(time_str):
                    try:
                        # Handle dashed time ranges like '18:00-19:00' by extracting start time
                        if '-' in str(time_str):
                            time_str = time_str.split('-')[0]
                        h, m = map(int, time_str.split(':'))
                        return h * 60 + m
                    except:
                        return -1

                start_minutes = time_to_minutes(original_start)
                end_minutes = time_to_minutes(original_end)

                print(f"🔍 周末工单合并计算: {work_order}")
                print(f"  原始时间: {original_start} -> {original_end}")
                print(f"  可用时间槽: {time_slots}")

                # 查找start和end对应的时间点索引
                start_idx = -1
                end_idx = -1

                # 查找开始时间点（精确匹配）
                for i, slot in enumerate(time_slots):
                    slot_minutes = time_to_minutes(slot)
                    if abs(slot_minutes - start_minutes) < 1:  # 使用<1避免浮点数问题
                        start_idx = i
                        print(f"  找到开始时间: {slot} (索引: {i})")
                        break

                # 查找结束时间点（精确匹配）
                for i, slot in enumerate(time_slots):
                    slot_minutes = time_to_minutes(slot)
                    if abs(slot_minutes - end_minutes) < 1:
                        end_idx = i
                        print(f"  找到结束时间: {slot} (索引: {i})")
                        break

                # 🔧 如果没找到精确匹配，尝试范围匹配
                if end_idx == -1:
                    print(f"⚠️ 未找到精确的结束时间 {original_end}，尝试范围匹配")
                    for i, slot in enumerate(time_slots):
                        slot_minutes = time_to_minutes(slot)
                        if start_minutes <= slot_minutes <= end_minutes:
                            end_idx = i
                            print(f"  范围匹配: {slot} (索引: {i})")

                    # 如果还是没找到，使用最接近但不超出的时间点
                    if end_idx == -1:
                        for i, slot in enumerate(time_slots):
                            slot_minutes = time_to_minutes(slot)
                            if slot_minutes <= end_minutes:
                                end_idx = i
                            else:
                                break

                        if end_idx >= 0:
                            print(f"  使用最接近时间点: {time_slots[end_idx]} (索引: {end_idx})")
                        else:
                            print(f"❌ 完全无法找到合适的结束时间点，使用最后一个时间槽")
                            end_idx = len(time_slots) - 1

                # 如果找到了开始时间
                if start_idx >= 0 and end_idx >= start_idx:
                    should_merge = True
                    start_slot = start_idx
                    end_slot = end_idx
                    print(f"✅ 最终合并范围: 索引{start_slot}->{end_slot}, 时间{time_slots[start_slot]}->{time_slots[end_slot]}")
                else:
                    print(f"❌ 无法确定合并范围，使用原始索引: {time_slot_index}")
                    if merge_info and merge_info.get('merge_cells'):
                        should_merge = True
                        start_slot = merge_info.get('start_slot_index', time_slot_index)
                        end_slot = merge_info.get('end_slot_index', time_slot_index)
            else:
                # 非周末或没有时间信息，使用原始逻辑
                if merge_info and merge_info.get('merge_cells'):
                    should_merge = True
                    start_slot = merge_info.get('start_slot_index', time_slot_index)
                    end_slot = merge_info.get('end_slot_index', time_slot_index)
                elif not is_weekend and original_start and original_end and time_slots:
                    # 工作日：如果时间范围与该日期的任何工单重叠，则放在第一列
                    # 工作日通常只有一个时间槽
                    if len(time_slots) > 0:
                        start_slot = 0
                        end_slot = 0

            if should_merge and start_slot < len(cols) and end_slot < len(cols):
                # 合并单元格
                min_col = cols[start_slot]
                max_col = cols[end_slot]
                ws.merge_cells(start_row=row_idx, start_column=min_col, end_row=row_idx, end_column=max_col)
                cell = ws.cell(row_idx, column=min_col)
                if work_order:
                    cell.value = work_order
                    cell.font = Font(name=font_family, size=12)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            else:
                # 不合并，直接放置
                if time_slot_index < len(cols):
                    target_col = cols[time_slot_index]

                    # 检查目标单元格是否已被合并
                    is_merged = False
                    for merged_cell in ws.merged_cells.ranges:
                        if (merged_cell.min_row <= row_idx <= merged_cell.max_row and
                            merged_cell.min_col <= target_col <= merged_cell.max_col):
                            is_merged = True
                            # 如果目标单元格在合并区域内，使用主单元格
                            row_idx = merged_cell.min_row
                            target_col = merged_cell.min_col
                            break

                    cell = ws.cell(row=row_idx, column=target_col)
                    try:
                        current_value = cell.value if cell.value else ""
                        if current_value and str(current_value).strip():
                            # 如果已经有内容，添加换行
                            cell.value = str(current_value) + '\n' + work_order
                        else:
                            cell.value = work_order
                        cell.font = Font(name=font_family, size=12)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                    except Exception as e:
                        print(f"⚠️ 警告：写入单元格 ({row_idx}, {target_col}) 时出错: {e}")
                        # 如果写入失败，尝试使用备用位置
                        try:
                            backup_cell = ws.cell(row=row_idx, column=cols[0])
                            backup_cell.value = work_order
                            backup_cell.font = Font(name=font_family, size=12)
                            backup_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            backup_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                        except:
                            print(f"❌ 错误：无法为工单找到合适的单元格位置")
                            raise

    # 合并周末的连续空白单元格
    for row_idx in range(4, len(personnel_assignments) + 4):  # 人员行从第4行开始
        for date in dates:
            if not date_is_weekend.get(date, False):  # 只处理周末
                continue

            cols = date_col_mapping.get(date, [])
            if len(cols) <= 1:  # 只有多列才需要合并空白单元格
                continue

            # 首先记录哪些列已经被工单占用（通过检查哪些单元格有值或被合并）
            occupied_cols = set()

            for i, col in enumerate(cols):
                try:
                    cell = ws.cell(row=row_idx, column=col)
                    # 检查单元格是否在合并区域内
                    cell_in_merged = False
                    for merged_cell in ws.merged_cells.ranges:
                        if (merged_cell.min_row <= row_idx <= merged_cell.max_row and
                            merged_cell.min_col <= col <= merged_cell.max_col):
                            # 单元格在合并区域内
                            # 只有主单元格（左上角）才有可读的值
                            if row_idx == merged_cell.min_row and col == merged_cell.min_col:
                                # 这是主单元格，可以读取值
                                if cell.value and str(cell.value).strip():
                                    cell_in_merged = True
                            else:
                                # 非主单元格，如果在合并区域内则认为被占用
                                cell_in_merged = True

                            if cell_in_merged:
                                # 标记整个合并区域占用的列
                                for merged_col in range(merged_cell.min_col, merged_cell.max_col + 1):
                                    if merged_col in cols:
                                        occupied_cols.add(cols.index(merged_col))
                            break

                    # 如果不在合并区域内，直接检查值
                    if not cell_in_merged:
                        if cell.value and str(cell.value).strip():
                            occupied_cols.add(i)

                except Exception as e:
                    # 如果读取单元格值时出错，跳过该单元格
                    print(f"⚠️ 警告：读取单元格 ({row_idx}, {col}) 时出错: {e}")
                    continue

            # 找到连续的空白单元格区域（未被占用的列）
            empty_start = -1
            empty_end = -1

            for i, col in enumerate(cols):
                if i in occupied_cols:
                    # 这一列已被占用，合并之前的空白区域
                    if empty_end > empty_start:
                        min_col = cols[empty_start]
                        max_col = cols[empty_end]
                        ws.merge_cells(start_row=row_idx, start_column=min_col, end_row=row_idx, end_column=max_col)
                    # 重置
                    empty_start = -1
                    empty_end = -1
                else:
                    # 这一列未被占用
                    if empty_start == -1:
                        empty_start = i
                        empty_end = i
                    else:
                        empty_end = i

            # 处理行末的空白区域
            if empty_start != -1 and empty_end > empty_start:
                min_col = cols[empty_start]
                max_col = cols[empty_end]
                ws.merge_cells(start_row=row_idx, start_column=min_col, end_row=row_idx, end_column=max_col)

    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=len(personnel_assignments)+3, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = thin_border

    # 然后专门为合并的空白单元格应用边框（确保左上角单元格有完整边框）
    for row_idx in range(4, len(personnel_assignments) + 4):
        for date in dates:
            if not date_is_weekend.get(date, False):
                continue
            cols = date_col_mapping.get(date, [])
            if len(cols) <= 1:
                continue

            # 重新检查空白区域并应用边框
            occupied_cols = set()

            for i, col in enumerate(cols):
                try:
                    cell = ws.cell(row=row_idx, column=col)
                    # 检查单元格是否在合并区域内
                    cell_in_merged = False
                    for merged_cell in ws.merged_cells.ranges:
                        if (merged_cell.min_row <= row_idx <= merged_cell.max_row and
                            merged_cell.min_col <= col <= merged_cell.max_col):
                            # 单元格在合并区域内
                            if row_idx == merged_cell.min_row and col == merged_cell.min_col:
                                # 主单元格，可以读取值
                                if cell.value and str(cell.value).strip():
                                    cell_in_merged = True
                            else:
                                # 非主单元格，如果在合并区域内则认为被占用
                                cell_in_merged = True

                            if cell_in_merged:
                                for merged_col in range(merged_cell.min_col, merged_cell.max_col + 1):
                                    if merged_col in cols:
                                        occupied_cols.add(cols.index(merged_col))
                            break

                    # 如果不在合并区域内，直接检查值
                    if not cell_in_merged:
                        if cell.value and str(cell.value).strip():
                            occupied_cols.add(i)

                except Exception as e:
                    # 如果读取单元格值时出错，跳过该单元格
                    print(f"⚠️ 警告：读取单元格 ({row_idx}, {col}) 时出错: {e}")
                    continue

            # 找到连续的空白单元格区域
            empty_start = -1
            empty_end = -1

            for i, col in enumerate(cols):
                if i in occupied_cols:
                    if empty_end > empty_start:
                        min_col = cols[empty_start]
                        max_col = cols[empty_end]
                        # 为合并后的空白单元格应用边框
                        cell = ws.cell(row=row_idx, column=min_col)
                        cell.border = thin_border
                    empty_start = -1
                    empty_end = -1
                else:
                    if empty_start == -1:
                        empty_start = i
                        empty_end = i
                    else:
                        empty_end = i

            # 处理行末的空白区域
            if empty_start != -1 and empty_end > empty_start:
                min_col = cols[empty_start]
                max_col = cols[empty_end]
                cell = ws.cell(row=row_idx, column=min_col)
                cell.border = thin_border


def create_schedule_sheet_from_final_format(ws, data):
    """
    处理assignments_final格式
    data: {
        'excel_structure': {
            'sheet2_structure': {
                'header_row_1': '应用變更人員時間安排表 2026-05-07',
                'row_2_dates': ['2026-05-04', '2026-05-05', ...],
                'row_3_times': ['18:00', '09:00', ...],
                'personnel_rows': ['李俊毅', '岑惠韜', ...],
                'colors': {...}
            }
        },
        'assignments_final': [
            {
                'personnel': '李俊毅',
                'bg_color': '#E8F5E8',
                'assignments': [
                    {
                        'date': '2026-05-07',
                        'time_slot': '18:00',
                        'work_order': '(中台) 02241 CCM 晏鵬禹',
                        'original_start': '18:00',
                        'original_end': '20:00',
                        'is_weekend': false
                    },
                    ...
                ],
                'total_count': 2
            },
            ...
        ]
    }
    """
    font_family = '新宋体'
    excel_structure = data.get('excel_structure', {})
    sheet2_structure = excel_structure.get('sheet2_structure', {})
    assignments_final = data.get('assignments_final', [])

    title = sheet2_structure.get('header_row_1', '应用變更人員時間安排表')
    row_2_dates = sheet2_structure.get('row_2_dates', [])
    row_3_times = sheet2_structure.get('row_3_times', [])
    personnel_rows = sheet2_structure.get('personnel_rows', [])
    colors = sheet2_structure.get('colors', {})

    # Issue 6: 标题格式修正
    title = format_title_date(title)

    # 创建列映射：每个列对应一个日期和时间
    # row_2_dates和row_3_times的长度应该相同
    num_cols = len(row_2_dates)
    current_col = 2

    # 分析每个日期是工作日还是周末
    date_is_weekend = {}
    for person_data in assignments_final:
        for assignment in person_data.get('assignments', []):
            date = assignment.get('date', '')
            is_weekend = assignment.get('is_weekend', False)
            if date and date not in date_is_weekend:
                date_is_weekend[date] = is_weekend

    # 为每个列设置列宽
    for i in range(num_cols):
        current_date = row_2_dates[i]
        # Issue 4: 根据工作日/周末设置不同列宽
        if date_is_weekend.get(current_date, False):
            ws.column_dimensions[get_column_letter(current_col)].width = 16  # 周末列宽16
        else:
            ws.column_dimensions[get_column_letter(current_col)].width = 32  # 工作日列宽32
        current_col += 1

    ws.column_dimensions['A'].width = 33
    total_cols = num_cols

    # 标题行
    ws.row_dimensions[1].height = 45
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(name=font_family, size=30, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_fill_color = fix_color_code('#E6F3FF')
    title_cell.fill = PatternFill(start_color=title_fill_color, end_color=title_fill_color, fill_type='solid')
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Issue 2: A2和A3合并，内容寫執行人員
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    person_label_cell = ws.cell(row=2, column=1)
    person_label_cell.value = '執行人員'
    person_label_cell.font = Font(name=font_family, size=14, bold=True)
    person_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    person_label_cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')

    # 日期行（合并相同的日期）
    ws.row_dimensions[2].height = 35
    date_fill_color = fix_color_code('#E6F3FF')

    i = 0
    while i < num_cols:
        current_date = row_2_dates[i]
        # 找到连续的相同日期
        j = i
        while j < num_cols and row_2_dates[j] == current_date:
            j += 1

        # i到j-1是相同的日期，合并它们
        start_c = i + 2  # +2因为列从2开始（A列是1）
        end_c = j + 1  # j-1+2
        cell = ws.cell(row=2, column=start_c)
        # Issue 3: 日期格式从MM-DD改为X月X日
        if '-' in current_date:
            date_parts = current_date.split('-')
            if len(date_parts) == 3:
                # 从 YYYY-MM-DD 格式转换为 X月X日
                month = date_parts[1].lstrip('0')
                day = date_parts[2].lstrip('0')
                display_date = f"{month}月{day}日"
            else:
                display_date = format_date_chinese(current_date)
        else:
            display_date = format_date_chinese(current_date)
        cell.value = display_date
        cell.font = Font(name=font_family, size=26, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color=date_fill_color, end_color=date_fill_color, fill_type='solid')
        if start_c != end_c:
            ws.merge_cells(start_row=2, start_column=start_c, end_row=2, end_column=end_c)

        i = j

    # 时间行
    ws.row_dimensions[3].height = 30
    for i, time_slot in enumerate(row_3_times):
        col = i + 2
        cell = ws.cell(row=3, column=col)
        cell.value = time_slot
        cell.font = Font(name=font_family, size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 人员行
    for row_idx, person_data in enumerate(assignments_final, start=4):
        person_name = person_data.get('personnel', '')
        bg_color_code = person_data.get('bg_color', '#FFFFFF')
        assignments = person_data.get('assignments', [])
        # Issue 5: 第4行以后的行行高80
        ws.row_dimensions[row_idx].height = 80
        bg_color = fix_color_code(bg_color_code)

        # 姓名单元格
        person_cell = ws.cell(row_idx, column=1)
        person_cell.value = person_name
        person_cell.font = Font(name=font_family, size=14, bold=True)
        person_cell.alignment = Alignment(horizontal='center', vertical='center')
        person_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        # 任务单元格
        for assignment in assignments:
            date = assignment.get('date', '')
            time_slot = assignment.get('time_slot', '')
            work_order = assignment.get('work_order', '')
            original_start = assignment.get('original_start', '')
            original_end = assignment.get('original_end', '')
            is_weekend = assignment.get('is_weekend', False)
            weekend_note = assignment.get('weekend_note', '')

            # 找到对应的列
            target_col = None
            # 首先尝试通过time_slot匹配
            for i, (row_date, row_time) in enumerate(zip(row_2_dates, row_3_times)):
                if row_date == date and row_time == time_slot:
                    target_col = i + 2
                    break

            # 如果没找到，尝试通过original_start匹配
            if target_col is None and original_start:
                for i, (row_date, row_time) in enumerate(zip(row_2_dates, row_3_times)):
                    if row_date == date and row_time == original_start:
                        target_col = i + 2
                        break

            # 如果还是没找到，尝试通过original_end匹配
            if target_col is None and original_end:
                for i, (row_date, row_time) in enumerate(zip(row_2_dates, row_3_times)):
                    if row_date == date and row_time == original_end:
                        target_col = i + 2
                        break

            # 如果还是找不到，输出警告并跳过
            if target_col is None:
                print(f"⚠ 警告：无法为工单找到对应的列")
                print(f"  人员: {person_name}")
                print(f"  工单: {work_order}")
                print(f"  日期: {date}, 时间槽: {time_slot}, 开始: {original_start}, 结束: {original_end}")
                continue

            if work_order:
                # 检查是否需要合并单元格
                should_merge = False
                merge_end_col = target_col

                if is_weekend and original_start and original_end:
                    # 周末工单，检查是否需要合并
                    # 查找original_end对应的时间列
                    for i, (row_date, row_time) in enumerate(zip(row_2_dates, row_3_times)):
                        if row_date == date and row_time == original_end:
                            merge_end_col = i + 2
                            if merge_end_col != target_col:
                                should_merge = True
                            break

                if should_merge and merge_end_col > target_col:
                    # 合并单元格
                    ws.merge_cells(start_row=row_idx, start_column=target_col, end_row=row_idx, end_column=merge_end_col)
                    cell = ws.cell(row_idx, column=target_col)
                    cell.value = work_order
                    cell.font = Font(name=font_family, size=12)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                else:
                    # 不合并，直接放置
                    cell = ws.cell(row_idx, column=target_col)
                    try:
                        current_value = cell.value if cell.value else ""
                        if current_value and str(current_value).strip():
                            # 如果已经有内容，添加换行
                            cell.value = str(current_value) + '\n' + work_order
                        else:
                            cell.value = work_order
                        cell.font = Font(name=font_family, size=12)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                    except Exception as e:
                        print(f"⚠️ 警告：写入单元格时出错: {e}")
                        # 如果写入失败，尝试直接覆盖
                        try:
                            cell.value = work_order
                            cell.font = Font(name=font_family, size=12)
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                        except:
                            print(f"❌ 错误：无法写入工单到单元格")
                            raise

    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=len(assignments_final)+3, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = thin_border


def enforce_balanced_assignment(data):
    """
    强制执行均衡分配规则：
    1. 每个人至少1个工单
    2. 每人工单数差异不超过1
    3. 如果AI分配不均衡，自动重新分配
    """
    import copy

    # 创建数据的深拷贝
    data_copy = copy.deepcopy(data)

    # 处理不同的格式
    if 'personnel_assignments' in data_copy:
        personnel_assignments = data_copy['personnel_assignments']

        # 收集所有工单
        all_work_orders = []
        for person in personnel_assignments:
            assignments = person.get('assignments', [])
            for assignment in assignments:
                all_work_orders.append({
                    'data': assignment,
                    'current_person': person.get('name', '')
                })

        total_work_orders = len(all_work_orders)
        total_people = len(personnel_assignments)

        if total_work_orders == 0 or total_people == 0:
            print("⚠ 警告：没有工单或人员，跳过均衡分配检查")
            return data_copy

        # 计算理想分配
        base_count = total_work_orders // total_people
        remainder = total_work_orders % total_people

        # 每人应有的工单数
        target_counts = {}
        for i, person in enumerate(personnel_assignments):
            target_counts[person.get('name', '')] = base_count + (1 if i < remainder else 0)

        print(f"=== 均衡分配检查 ===")
        print(f"总工单数: {total_work_orders}, 人员数: {total_people}")
        print(f"理想分配: 每人{base_count}个工单，前{remainder}人多1个")

        # 检查当前分配是否均衡
        current_counts = {}
        for person in personnel_assignments:
            name = person.get('name', '')
            count = len(person.get('assignments', []))
            current_counts[name] = count

        print(f"当前分配: {current_counts}")

        # 检查是否有人0工单
        zero_count_people = [name for name, count in current_counts.items() if count == 0]
        if zero_count_people:
            print(f"❌ 致命错误：以下人员有0工单: {zero_count_people}")
            print("开始自动重新分配...")

            # 清空所有人的分配
            for person in personnel_assignments:
                person['assignments'] = []
                person['total_count'] = 0

            # 按照目标分配重新分配工单
            person_idx = 0
            for work_order in all_work_orders:
                # 找到下一个还需要工单的人
                while person_idx < len(personnel_assignments):
                    person = personnel_assignments[person_idx]
                    person_name = person.get('name', '')
                    current_count = len(person.get('assignments', []))
                    target_count = target_counts[person_name]

                    if current_count < target_count:
                        # 分配工单给这个人
                        assignment = work_order['data'].copy()
                        person['assignments'].append(assignment)
                        person['total_count'] = len(person['assignments'])
                        break

                    person_idx += 1

                # 如果所有人都达到了目标，从头开始
                if person_idx >= len(personnel_assignments):
                    person_idx = 0

            # 验证重新分配的结果
            print(f"重新分配后的结果:")
            for person in personnel_assignments:
                name = person.get('name', '')
                count = len(person.get('assignments', []))
                print(f"  {name}: {count}个工单")

        # 检查是否差异超过1
        counts = list(current_counts.values())
        if max(counts) - min(counts) > 1:
            print(f"❌ 警告：工单分配不均衡，差异={max(counts) - min(counts)}")
            # 这里也可以添加重新分配逻辑

    return data_copy




# ====================== 完整还原：生成 2 个工作表 ======================
def create_excel_from_json(data, output_file='output_transformed.xlsx'):
    # 首先执行均衡分配检查和自动修正（最高优先级）
    data = enforce_balanced_assignment(data)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 工作表1：排班表（第一个工作表）
    try:
        if 'excel_structure' in data and 'assignments_final' in data:
            # 最新格式：excel_structure + assignments_final
            ws1 = wb.create_sheet("變更人員時間安排表", 0)
            create_schedule_sheet_from_final_format(ws1, data)
        elif 'schedule_structure' in data and 'personnel_assignments' in data:
            # schedule_structure + personnel_assignments格式
            ws1 = wb.create_sheet("變更人員時間安排表", 0)
            create_schedule_sheet_from_new_structure_format(ws1, data)
        elif 'structure' in data:
            structure = data['structure']
            # 检查structure的子格式
            if 'dates' in structure and 'staff_list' in structure:
                # 旧格式：有dates和staff_list字段
                ws1 = wb.create_sheet("變更人員時間安排表", 0)
                create_schedule_sheet_from_latest_format(ws1, structure.get('header', '应用變更人員時間安排表'), structure)
            elif 'headers' in structure and 'personnel_rows' in structure:
                # 新格式：有headers和personnel_rows字段
                ws1 = wb.create_sheet("變更人員時間安排表", 0)
                create_schedule_sheet_from_headers_format(ws1, structure)
            else:
                print("⚠ structure格式不识别，跳过排班表生成")
        elif 'schedule_data' in data:
            ws1 = wb.create_sheet("變更人員時間安排表", 0)
            create_schedule_sheet_from_new_format(ws1, data.get('header', '应用變更人員時間安排表'), data.get('participants', []), data)
        elif 'sheet_structure' in data and 'assignments' in data:
            # assignments格式：有sheet_structure和assignments字段
            ws1 = wb.create_sheet("變更人員時間安排表", 0)
            create_schedule_sheet_from_assignments_format(ws1, data)
        else:
            print(f"⚠ 未识别的JSON格式，可用字段: {list(data.keys())}")
            ws1 = wb.create_sheet("變更人員時間安排表", 0)
    except Exception as e:
        print(f"排班表生成异常: {e}")
        import traceback
        traceback.print_exc()

    # 工作表2：变更明细表（第二个工作表）
    df = load_filtered_excel_data()
    if df is not None:
        ws2 = wb.create_sheet("變更明細數據報表", 1)
        create_data_sheet_from_dataframe(ws2, df)

    wb.save(output_file)
    return output_file


def main():
    print("JSON → Excel 转换中...")
    data = load_json_result()
    create_excel_from_json(data)
    print("✅ 完成！双工作表已生成")

if __name__ == "__main__":
    main()